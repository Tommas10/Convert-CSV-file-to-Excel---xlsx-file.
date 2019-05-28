#!/usr/bin/env python

#Small automation Python script- Convert CSV file to Excel - xlsx file.
#Created by Tommas Huang 
#Created date: 2019-05-29

import os
#Python OS module provides easy functions that allow us to interact 
#and get Operating System related information and even control processes up to a limit.
import csv
#The csv module implements classes to read and write tabular data in CSV format.
import sys
#This module provides access to some variables used or maintained by the interpreter and to functions that interact strongly with the interpreter.
from openpyxl import Workbook
#Create Excel sheet

if __name__ == '__main__':
#it sets a few special variables like __name__, and then
#it executes all of the code found in the file.
    workbook = Workbook()
    #The Workbook() constructor is used to create a new Excel workbook with a given filename
    worksheet = workbook.active
    #A workbook is always created with at least one worksheet.
    with open('/Users/tommashuang/Downloads/100SalesRecords.csv', 'r') as f:
    #Open CSV file from source.
        reader = csv.reader(f)
        #Return a reader object which will iterate over lines in the given csvfile. 
        for r, row in enumerate(reader):
            for c, col in enumerate(row):
                for idx, val in enumerate(col.split(',')):
                    cell = worksheet.cell(row=r+1, column=c+1)
                    cell.value = val
    workbook.save('output.xlsx')
    #Save convert CSV file to xlsx file